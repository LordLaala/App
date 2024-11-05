# -*- coding: utf-8 -*-

import os
import json
from dateutil.parser import parse
from Logger import Logger
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
class DataframeMaker:
    def __init__(self, json_dir):
        self.json_dir = json_dir

    @staticmethod
    def is_date(string):
        try:
            parse(string, fuzzy=True, dayfirst=True)
            return True
        except ValueError:
            return False

    def make_json_to_df(self):
        log = Logger(path_name="log_dc.txt")
        input_file = self.json_dir
        merged_file = {"Gemeinde":[],"Zuteilung":[], "Kat.-Nr. (aktuell)":[],"Kat.-Nr. (alt)":[],"Eintragsdatum":[],"Ereignis":[],"Zusaetzliche Informationen":[]}
        for dir, root, files in os.walk(self.json_dir):
            print(len(files))
            for file in files:
                try:
                    with open(os.path.join(input_file, file),encoding='utf-8',mode='r') as json_file:
                        datum = [0]
                        ereignis = [0]
                        is_first_date = True
                        did_the_change = False
                        try:
                            data = json.load(json_file)
                        except Exception as e:
                            print(file)
                            continue
                        if len(data) <= 1:
                            log.write_error_json_ZI(file)
                            continue
                        for keys in data:
                            for x in data[keys]:
                                if x in ['Gemeinde', 'Kat.-Nr. (aktuell)', 'Zuteilung',
                                                             'Kat.-Nr. (alt)', 'Datum',
                                                             'Datum:', 'Alt Kat. Nr.']:
                                    merged_file[x].append(keys)
                                    did_the_change = True
                                if x in ["Kat.-Nr. (alt) Ereignis"]:
                                    merged_file["Kat.-Nr. (alt)"].append(keys)
                                    did_the_change = True
                            if did_the_change:
                                did_the_change = False
                                continue
                            match keys:

                                case keys if DataframeMaker.is_date(keys):
                                    if is_first_date:
                                        datum = []
                                        ereignis = []
                                        is_first_date = False
                                    datum.append(keys)
                                    for x in data[keys]:
                                        ereignis.append(x)

                                case keys if keys in ["Kat.-Nr. (alt)","Alt Kat. Nr."] or "alt" in keys.lower():
                                    merged_file['Kat.-Nr. (alt)'].append(data[keys][0])


                                case keys if "alt" in keys.lower():
                                    print(file)

                                #case keys if keys == "Zusaetzliche Informationen":
                                #   if data[keys][0] == '' and len(data[keys]) == 1:
                                #        merged_file["Zusaetzliche Informationen"].append(['False'])
                                #    else:
                                #        Z_I_List = []
                                #        merged_file["Zusaetzliche Informationen"].append(['True'])
                                #        Z_I_List.append(data[keys])

                                case keys if keys in list(merged_file.keys()):
                                        if len(data[keys]) == 0:
                                            merged_file[keys].append('None')
                                        else:
                                            merged_file[keys].append(data[keys][0])
                                case _:
                                    continue
                        length = len(merged_file['Gemeinde'])

                        for keys in merged_file:
                            if len(merged_file[keys]) != length:
                                if keys not in ["Eintragsdatum","Ereignis"]:
                                    merged_file[keys].append('Not Titled')


                        merged_file["Ereignis"].append(ereignis)
                        merged_file["Eintragsdatum"].append(datum)
                except Exception as e:
                    print(f'{e}fdgfd :{file}')
                    continue
        df = pd.DataFrame.from_dict(merged_file)
        """
    
        df_help = df.copy()
        for x in range(0,len(df["Zusaetzliche Informationen"])):
            if df["Zusaetzliche Informationen"][x] == "":
                df["Zusaetzliche Informationen"][x] = "False"
            else:
                df["Zusaetzliche Informationen"][x] = "True"

        for x in range(0, len(df_help["Zusaetzliche Informationen"])):
            if df_help["Zusaetzliche Informationen"][x] == "":
                df_help["Zusaetzliche Informationen"][x] = np.nan
        df_help.dropna(inplace=True)
        df_help.drop(columns=["Eintragsdatum","Ereignis"],inplace=True)
"""

        return df
def is_date_fn(string):
    try:
        parse(string, fuzzy=True, dayfirst=True)
        return True
    except ValueError:
        return False

def mark_invalid_date(value):
    if is_date_fn(value) and len(value)>5:
        return value
    else:
        return f'{value}[Invalid]'

def expand_and_clean_df(df):
    expanded_df = pd.DataFrame({
        'Gemeinde': df['Gemeinde'].repeat(df['Eintragsdatum'].apply(len)).values,
        'Zuteilung': df['Zuteilung'].repeat(df['Eintragsdatum'].apply(len)).values,
        'Kat.-Nr. (aktuell)': df['Kat.-Nr. (aktuell)'].repeat(df['Eintragsdatum'].apply(len)).values,
        'Kat.-Nr. (alt)': df['Kat.-Nr. (alt)'].repeat(df['Eintragsdatum'].apply(len)).values,
        'Eintragsdatum': df['Eintragsdatum'].explode().values,
        'Ereignis': df['Ereignis'].explode().values,
        'Zusaetzliche Informationen': df['Zusaetzliche Informationen'].repeat(df['Eintragsdatum'].apply(len)).values
    })
    for x in range(len(expanded_df["Eintragsdatum"])):
        if expanded_df.loc[x, "Eintragsdatum"] == 0:
            continue
        else:
            expanded_df.loc[x, "Eintragsdatum"] = mark_invalid_date(expanded_df.loc[x, "Eintragsdatum"])
    return expanded_df

def save_dataframe_to_excel(dataframe, excel_file_path):
    os.makedirs(os.path.dirname(excel_file_path), exist_ok=True)

    with pd.ExcelWriter(excel_file_path, engine="openpyxl") as writer:
        dataframe.to_excel(
            writer, index=False, sheet_name="Sheet1", startrow=1, startcol=1
        )
        writer._save()

    # Load the workbook and the sheet
    workbook = load_workbook(excel_file_path)

    sheet = workbook["Sheet1"]

    # Create a table
    table = Table(displayName="Table1", ref=sheet.dimensions)

    # Add a table style
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=True,
    )
    table.tableStyleInfo = style

    # Add the table to the sheet
    sheet.add_table(table)

    # Save the workbook
    workbook.save(excel_file_path)

def adjust_excel_table(excel_file_path):
    # Load the workbook and the sheet
    workbook = load_workbook(excel_file_path)
    sheet = workbook["Sheet1"]

    # Adjust the width of each column to fit the content
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length) + 5
        sheet.column_dimensions[column].width = adjusted_width

    # Adjust the height of each row to fit the content
    for row in sheet.iter_rows():
        max_height = 0
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)
            if cell.value:
                lines = str(cell.value).split("\n")
                max_height = max(max_height, len(lines))
        sheet.row_dimensions[row[0].row].height = (
            max_height * 15
        )  # Approximate height per line

    # Save the workbook
    workbook.save(excel_file_path)

def main(output_JSON,output_Excel):
    excel = DataframeMaker(output_JSON)
    df = excel.make_json_to_df()
    df = expand_and_clean_df(df)
    save_dataframe_to_excel(df, os.path.join(output_Excel,"Journal.xlsx"))
    adjust_excel_table(os.path.join(output_Excel,"Journal.xlsx"))

main("JSON","C:\\Users\\BAA3012\PycharmProjects\pythonProject1\\Excel")